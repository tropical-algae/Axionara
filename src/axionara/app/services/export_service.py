import csv
import io
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException, Response
from openpyxl import load_workbook
from sqlmodel import Session

from axionara.app.services.storage_service import get_storage_service
from axionara.app.utils.constant import CONSTANT
from axionara.common.config import settings
from axionara.common.util import generate_random_token
from axionara.core.db.crud import (
    insert_export_job,
    select_active_access_grant,
    select_dataset_by_id,
    select_dataset_profile_by_dataset_id,
    select_export_job_by_id,
    select_export_jobs_by_user,
    select_export_jobs_by_user_dataset,
    update_export_job,
)
from axionara.core.db.models import DatasetAsset, ExportJob, UserAccount
from axionara.core.db.session import local_session
from axionara.core.model.dataset import DatasetAssetStatus, ExportJobStatus


class ExportService:
    def __init__(self):
        self.storage = get_storage_service()

    def create_export_job(
        self, db: Session, dataset_id: str, target_format: str, user: UserAccount
    ) -> ExportJob:
        dataset = self._get_published_dataset(db=db, dataset_id=dataset_id)
        grant = select_active_access_grant(db=db, dataset_id=dataset_id, user_id=user.id)
        if grant is None:
            raise HTTPException(**CONSTANT.RESP_ACCESS_GRANT_NOT_EXISTS)

        profile = select_dataset_profile_by_dataset_id(db=db, dataset_id=dataset_id)
        allowed_formats = profile.allowed_export_formats if profile is not None else []
        if target_format not in allowed_formats:
            raise HTTPException(**CONSTANT.RESP_EXPORT_FORMAT_UNSUPPORTED)

        return insert_export_job(
            db=db,
            job=ExportJob(
                id=generate_random_token(prefix="EXP", length=24),
                dataset_id=dataset.id,
                user_id=user.id,
                grant_id=grant.id,
                target_format=target_format,
                job_status=ExportJobStatus.PENDING.value,
            ),
        )

    def process_export_job(self, db: Session, job_id: str) -> ExportJob:
        job = self._get_export_job(db=db, job_id=job_id)
        dataset = self._get_published_dataset(db=db, dataset_id=job.dataset_id)

        job.job_status = ExportJobStatus.RUNNING.value
        job.started_at = datetime.now()
        update_export_job(db=db, job=job)

        try:
            artifact = self._build_export_artifact(
                dataset=dataset, target_format=job.target_format
            )
            if job.target_format == "raw":
                job.output_bucket = dataset.raw_bucket
                job.output_object_key = dataset.raw_object_key
            else:
                object_key = self._artifact_object_key(
                    dataset=dataset, job=job, filename=artifact["filename"]
                )
                stored = self.storage.save_bytes(
                    bucket=settings.MINIO_BUCKET_ARTIFACTS,
                    object_key=object_key,
                    content=artifact["content"],
                    content_type=artifact["content_type"],
                )
                job.output_bucket = stored.bucket
                job.output_object_key = stored.object_key

            job.output_filename = artifact["filename"]
            job.output_content_type = artifact["content_type"]
            job.output_size_bytes = len(artifact["content"])
            job.job_status = ExportJobStatus.SUCCEEDED.value
            job.finished_at = datetime.now()
            return update_export_job(db=db, job=job)
        except Exception as err:
            job.job_status = ExportJobStatus.FAILED.value
            job.error_message = str(err)
            job.finished_at = datetime.now()
            update_export_job(db=db, job=job)
            raise

    def list_my_export_jobs(
        self, db: Session, user: UserAccount, dataset_id: str | None = None
    ) -> list[ExportJob]:
        if dataset_id is not None:
            return select_export_jobs_by_user_dataset(
                db=db, user_id=user.id, dataset_id=dataset_id
            )
        return select_export_jobs_by_user(db=db, user_id=user.id)

    def get_my_export_job(self, db: Session, job_id: str, user: UserAccount) -> ExportJob:
        job = self._get_export_job(db=db, job_id=job_id)
        if job.user_id != user.id:
            raise HTTPException(**CONSTANT.RESP_DATASET_FORBIDDEN)
        return job

    def retry_export_job(self, db: Session, job_id: str, user: UserAccount) -> ExportJob:
        job = self.get_my_export_job(db=db, job_id=job_id, user=user)
        if job.job_status != ExportJobStatus.FAILED.value:
            raise HTTPException(**CONSTANT.RESP_EXPORT_JOB_NOT_RETRYABLE)
        return self.create_export_job(
            db=db,
            dataset_id=job.dataset_id,
            target_format=job.target_format,
            user=user,
        )

    def download_export(self, db: Session, job_id: str, user: UserAccount) -> Response:
        job = self.get_my_export_job(db=db, job_id=job_id, user=user)
        if (
            job.job_status != ExportJobStatus.SUCCEEDED.value
            or job.output_bucket is None
            or job.output_object_key is None
        ):
            raise HTTPException(**CONSTANT.RESP_EXPORT_JOB_NOT_READY)

        content = self.storage.get_bytes(
            bucket=job.output_bucket,
            object_key=job.output_object_key,
        )
        headers = {
            "Content-Disposition": (
                f'attachment; filename="{job.output_filename or job.id}"'
            )
        }
        return Response(
            content=content,
            media_type=job.output_content_type or "application/octet-stream",
            headers=headers,
        )

    def _build_export_artifact(
        self, dataset: DatasetAsset, target_format: str
    ) -> dict[str, Any]:
        raw_content = self.storage.get_bytes(
            bucket=dataset.raw_bucket or "",
            object_key=dataset.raw_object_key or "",
        )
        base_name = self._safe_base_filename(dataset)
        if target_format == "raw":
            return {
                "filename": dataset.original_filename,
                "content": raw_content,
                "content_type": dataset.content_type or "application/octet-stream",
            }
        if target_format == "json" and dataset.source_format == "json":
            payload = json.loads(raw_content.decode("utf-8-sig", errors="replace"))
            return self._json_artifact(base_name=base_name, payload=payload)

        rows = self._load_tabular_rows(dataset=dataset, raw_content=raw_content)
        if target_format == "csv":
            return self._csv_artifact(base_name=base_name, rows=rows)
        if target_format == "json":
            return self._json_artifact(base_name=base_name, payload=rows)
        if target_format == "sql":
            return self._sql_artifact(base_name=base_name, rows=rows)
        raise HTTPException(**CONSTANT.RESP_EXPORT_FORMAT_UNSUPPORTED)

    def _load_tabular_rows(
        self, dataset: DatasetAsset, raw_content: bytes
    ) -> list[dict[str, Any]]:
        if dataset.source_format == "csv":
            text = raw_content.decode("utf-8-sig", errors="replace")
            return list(csv.DictReader(io.StringIO(text)))
        if dataset.source_format == "json":
            text = raw_content.decode("utf-8-sig", errors="replace")
            payload = json.loads(text)
            if isinstance(payload, list) and all(
                isinstance(row, dict) for row in payload
            ):
                return payload
        if dataset.source_format == "xlsx":
            return self._load_xlsx_rows(raw_content=raw_content)
        raise HTTPException(**CONSTANT.RESP_EXPORT_FORMAT_UNSUPPORTED)

    def _load_xlsx_rows(self, raw_content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(
            filename=io.BytesIO(raw_content),
            read_only=True,
            data_only=True,
        )
        for worksheet in workbook.worksheets:
            rows = [
                list(row)
                for row in worksheet.iter_rows(values_only=True)
                if any(cell is not None for cell in row)
            ]
            if not rows:
                continue
            columns = [
                str(value).strip() if value not in (None, "") else f"column_{index + 1}"
                for index, value in enumerate(rows[0])
            ]
            return [
                {columns[index]: value for index, value in enumerate(row[: len(columns)])}
                for row in rows[1:]
            ]
        return []

    def _csv_artifact(self, base_name: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
        output = io.StringIO()
        fieldnames = self._fieldnames(rows=rows)
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
        return {
            "filename": f"{base_name}.csv",
            "content": output.getvalue().encode("utf-8"),
            "content_type": "text/csv; charset=utf-8",
        }

    def _json_artifact(self, base_name: str, payload: Any) -> dict[str, Any]:
        return {
            "filename": f"{base_name}.json",
            "content": json.dumps(
                payload, ensure_ascii=False, indent=2, default=str
            ).encode("utf-8"),
            "content_type": "application/json; charset=utf-8",
        }

    def _sql_artifact(self, base_name: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
        table_name = self._safe_identifier(base_name)
        columns = self._fieldnames(rows=rows)
        lines = [f"CREATE TABLE `{table_name}` ("]
        lines.extend(f"  `{self._safe_identifier(column)}` TEXT," for column in columns)
        if lines[-1].endswith(","):
            lines[-1] = lines[-1].rstrip(",")
        lines.append(");")
        for row in rows:
            column_sql = ", ".join(
                f"`{self._safe_identifier(column)}`" for column in columns
            )
            value_sql = ", ".join(
                self._sql_literal(row.get(column)) for column in columns
            )
            lines.append(
                f"INSERT INTO `{table_name}` ({column_sql}) VALUES ({value_sql});"
            )
        return {
            "filename": f"{base_name}.sql",
            "content": ("\n".join(lines) + "\n").encode("utf-8"),
            "content_type": "application/sql; charset=utf-8",
        }

    def _fieldnames(self, rows: list[dict[str, Any]]) -> list[str]:
        columns = []
        seen = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    columns.append(str(key))
                    seen.add(key)
        return columns or ["value"]

    def _sql_literal(self, value: Any) -> str:
        if value is None:
            return "NULL"
        return "'" + str(value).replace("'", "''") + "'"

    def _safe_base_filename(self, dataset: DatasetAsset) -> str:
        stem = Path(dataset.original_filename).stem or dataset.title or dataset.id
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", stem).strip("._")
        return safe or dataset.id

    def _safe_identifier(self, value: str) -> str:
        safe = re.sub(r"[^A-Za-z0-9_]+", "_", value).strip("_").lower()
        return safe or "dataset"

    def _artifact_object_key(
        self, dataset: DatasetAsset, job: ExportJob, filename: str
    ) -> str:
        return f"exports/{dataset.id}/{job.id}/{filename}"

    def _get_export_job(self, db: Session, job_id: str) -> ExportJob:
        job = select_export_job_by_id(db=db, job_id=job_id)
        if job is None:
            raise HTTPException(**CONSTANT.RESP_EXPORT_JOB_NOT_EXISTS)
        return job

    def _get_published_dataset(self, db: Session, dataset_id: str) -> DatasetAsset:
        dataset = select_dataset_by_id(db=db, dataset_id=dataset_id)
        if dataset is None or dataset.status != DatasetAssetStatus.PUBLISHED.value:
            raise HTTPException(**CONSTANT.RESP_DATASET_NOT_EXISTS)
        return dataset


def process_export_job_background(job_id: str) -> None:
    with local_session() as db:
        ExportService().process_export_job(db=db, job_id=job_id)
