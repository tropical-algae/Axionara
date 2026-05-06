import csv
import io
import json

from openpyxl import load_workbook

from axionara.core.db.models import DatasetAsset
from axionara.core.processing.parsers.base import BaseParser
from axionara.core.processing.types import ParsedResult


class CsvParser(BaseParser):
    def parse(self, dataset: DatasetAsset, content: bytes) -> ParsedResult:
        _ = dataset
        text = content.decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        rows = list(reader)
        columns = reader.fieldnames or []
        return ParsedResult(
            representation_type="tabular",
            schema_snapshot={"columns": [{"name": name} for name in columns]},
            data=rows,
            preview_data=rows[:10],
            parser_notes=[f"parsed_rows={len(rows)}"],
        )


class JsonParser(BaseParser):
    def parse(self, dataset: DatasetAsset, content: bytes) -> ParsedResult:
        _ = dataset
        payload = json.loads(content.decode("utf-8-sig", errors="replace"))
        if isinstance(payload, list) and all(isinstance(item, dict) for item in payload):
            columns = sorted({key for row in payload for key in row})
            return ParsedResult(
                representation_type="tabular",
                schema_snapshot={"columns": [{"name": name} for name in columns]},
                data=payload,
                preview_data=payload[:10],
                parser_notes=[f"parsed_rows={len(payload)}"],
            )
        return ParsedResult(
            representation_type="hierarchical",
            schema_snapshot={"root_type": type(payload).__name__},
            data=payload,
            preview_data=payload,
        )


class XlsxParser(BaseParser):
    def parse(self, dataset: DatasetAsset, content: bytes) -> ParsedResult:
        _ = dataset
        workbook = load_workbook(
            filename=io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
        for worksheet in workbook.worksheets:
            rows = [
                list(row)
                for row in worksheet.iter_rows(values_only=True)
                if any(cell is not None for cell in row)
            ]
            if not rows:
                continue
            columns = [
                str(value).strip() if value not in (None, "") else f"column_{index + 1}"
                for index, value in enumerate(rows[0])
            ]
            records = [
                {columns[index]: value for index, value in enumerate(row[: len(columns)])}
                for row in rows[1:]
            ]
            return ParsedResult(
                representation_type="tabular",
                schema_snapshot={
                    "sheet_name": worksheet.title,
                    "columns": [{"name": name} for name in columns],
                },
                data=records,
                preview_data=records[:10],
                parser_notes=[
                    f"parsed_rows={len(records)}",
                    f"sheet_name={worksheet.title}",
                ],
            )
        return ParsedResult(
            representation_type="tabular",
            parser_status="skipped",
            schema_snapshot={"columns": []},
            parser_notes=["xlsx workbook has no non-empty sheet"],
        )


class TextParser(BaseParser):
    def parse(self, dataset: DatasetAsset, content: bytes) -> ParsedResult:
        text = content.decode("utf-8-sig", errors="replace")
        return ParsedResult(
            representation_type="document",
            schema_snapshot={"document_type": dataset.source_format},
            extracted_text=text[:4000],
            parser_notes=[f"extractable_text_chars={len(text)}"],
        )


class PdfParser(BaseParser):
    def parse(self, dataset: DatasetAsset, content: bytes) -> ParsedResult:
        _ = dataset, content
        return ParsedResult(
            representation_type="document",
            schema_snapshot={"document_type": "pdf"},
            parser_notes=["pdf text extraction pending"],
        )


def get_parser(source_format: str) -> BaseParser:
    parser_map: dict[str, type[BaseParser]] = {
        "csv": CsvParser,
        "json": JsonParser,
        "xlsx": XlsxParser,
        "txt": TextParser,
        "pdf": PdfParser,
    }
    return parser_map[source_format]()
